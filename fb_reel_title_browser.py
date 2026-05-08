#!/usr/bin/env python3

import argparse
import sys
import re
import html
from datetime import datetime, timezone
from pathlib import Path
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright


DEFAULT_XLSX = Path("shadi_shirri_reels.xlsx")
REEL_URL_COLUMN = "reel_url"
REEL_TITLE_COLUMN = "reel_title"
UPLOAD_DATE_COLUMN = "upload_date"


def clean_text(text: str) -> str:
    text = html.unescape(text or "")
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\s*\|\s*Facebook$", "", text, flags=re.I)
    text = re.sub(r"\s+on Facebook$", "", text, flags=re.I)

    parts = [p.strip() for p in text.split("|") if p.strip()]
    if len(parts) >= 2:
        useful_parts = [
            p
            for p in parts
            if not re.search(r"\b(views?|reactions?|comments?|shares?)\b", p, flags=re.I)
            and p.lower() not in {"facebook", "shadi shirri"}
        ]
        if useful_parts:
            text = useful_parts[0]

    return text.strip()


def extract_title_from_html(page_html: str, page_title: str) -> str | None:
    soup = BeautifulSoup(page_html, "html.parser")

    candidates = []

    for attr_name in ["og:title", "twitter:title", "description", "og:description"]:
        tag = soup.find("meta", attrs={"property": attr_name}) or soup.find(
            "meta", attrs={"name": attr_name}
        )
        if tag and tag.get("content"):
            candidates.append(tag["content"])

    candidates.append(page_title)

    for candidate in candidates:
        title = clean_text(candidate)
        if title and title.lower() not in {
            "facebook",
            "facebook reels",
            "log into facebook",
            "watch",
        }:
            return title

    return None


def timestamp_to_date(ts: str | int | float) -> str | None:
    try:
        ts = int(float(ts))

        # Handle milliseconds
        if ts > 10_000_000_000:
            ts = ts // 1000

        dt = datetime.fromtimestamp(ts, tz=timezone.utc)
        return dt.strftime("%Y-%m-%d %H:%M:%S UTC")
    except Exception:
        return None


def extract_date_from_html(page_html: str) -> str | None:
    soup = BeautifulSoup(page_html, "html.parser")

    # 1. Normal HTML <time datetime="...">
    time_tag = soup.find("time")
    if time_tag:
        if time_tag.get("datetime"):
            return clean_text(time_tag["datetime"])
        if time_tag.text:
            return clean_text(time_tag.text)

    # 2. Common JSON keys Facebook sometimes embeds
    patterns = [
        r'"publish_time"\s*:\s*(\d+)',
        r'"creation_time"\s*:\s*(\d+)',
        r'"created_time"\s*:\s*(\d+)',
        r'"creation_timestamp"\s*:\s*(\d+)',
        r'"timestamp"\s*:\s*(\d{10,13})',
        r'"creation_time"\s*:\s*\{"timestamp"\s*:\s*(\d+)',
        r'"publish_time"\s*:\s*\{"timestamp"\s*:\s*(\d+)',
    ]

    for pattern in patterns:
        match = re.search(pattern, page_html)
        if match:
            date = timestamp_to_date(match.group(1))
            if date:
                return date

    # 3. ISO-like dates embedded in scripts
    iso_patterns = [
        r'"datePublished"\s*:\s*"([^"]+)"',
        r'"uploadDate"\s*:\s*"([^"]+)"',
        r'"created_time"\s*:\s*"([^"]+)"',
    ]

    for pattern in iso_patterns:
        match = re.search(pattern, page_html)
        if match:
            return clean_text(match.group(1))

    return None


def get_reel_info_from_page(page, url: str) -> tuple[str, str]:
    page.goto(url, wait_until="domcontentloaded", timeout=45000)
    page.wait_for_timeout(6000)

    page_html = page.content()
    page_title = page.title()

    title = extract_title_from_html(page_html, page_title)
    date = extract_date_from_html(page_html)

    if not title:
        title = ""

    if not date:
        date = ""

    return title, date


def get_reel_info(url: str, *, headless: bool = True) -> tuple[str, str]:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        page = new_page(browser)
        try:
            return get_reel_info_from_page(page, url)
        finally:
            browser.close()


def new_page(browser):
    return browser.new_page(
        locale="en-US",
        user_agent=(
            "Mozilla/5.0 (X11; Linux x86_64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/121.0 Safari/537.36"
        ),
    )


def norm_cell(value) -> str:
    return str(value or "").strip()


def ensure_columns(ws, headers: list[str], wanted: list[str]) -> dict[str, int]:
    header_to_col = {h: i + 1 for i, h in enumerate(headers) if h}
    for name in wanted:
        if name not in header_to_col:
            col = ws.max_column + 1
            ws.cell(row=1, column=col).value = name
            header_to_col[name] = col
    return header_to_col


def should_write(value: str, *, force: bool) -> bool:
    return force or not norm_cell(value)


def update_excel_reel_info(
    *,
    xlsx_path: Path,
    limit: int | None = None,
    start_row: int = 2,
    force: bool = False,
    headless: bool = True,
    save_every: int = 10,
) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    headers = [norm_cell(c.value) for c in ws[1]]
    cols = ensure_columns(ws, headers, [REEL_URL_COLUMN, REEL_TITLE_COLUMN, UPLOAD_DATE_COLUMN])

    url_col = cols[REEL_URL_COLUMN]
    title_col = cols[REEL_TITLE_COLUMN]
    date_col = cols[UPLOAD_DATE_COLUMN]

    processed = 0
    changed = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        page = new_page(browser)
        try:
            for row_idx in range(max(2, start_row), ws.max_row + 1):
                if limit is not None and processed >= limit:
                    break

                reel_url = norm_cell(ws.cell(row=row_idx, column=url_col).value)
                if not reel_url:
                    continue

                current_title = norm_cell(ws.cell(row=row_idx, column=title_col).value)
                current_date = norm_cell(ws.cell(row=row_idx, column=date_col).value)
                if not should_write(current_title, force=force) and not should_write(current_date, force=force):
                    continue

                processed += 1
                print(f"[{processed}] row={row_idx} fetching {reel_url}", flush=True)

                try:
                    title, date = get_reel_info_from_page(page, reel_url)
                except Exception as e:
                    print(f"    ERROR: {e}", flush=True)
                    continue

                row_changed = False
                if title and should_write(current_title, force=force):
                    ws.cell(row=row_idx, column=title_col).value = title
                    row_changed = True
                    print(f"    title: {title}", flush=True)
                else:
                    print("    title: not found", flush=True)

                if date and should_write(current_date, force=force):
                    ws.cell(row=row_idx, column=date_col).value = date
                    row_changed = True
                    print(f"    date: {date}", flush=True)
                else:
                    print("    date: not found", flush=True)

                if row_changed:
                    changed += 1

                if save_every > 0 and changed > 0 and changed % save_every == 0:
                    wb.save(xlsx_path)
                    print(f"    saved {xlsx_path}", flush=True)
        finally:
            browser.close()

    wb.save(xlsx_path)
    print(f"Done. Processed={processed}, changed_rows={changed}, file={xlsx_path}", flush=True)
    return changed


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract Facebook reel titles/dates and optionally write them into the Excel file."
    )
    parser.add_argument("url", nargs="?", help="Single Facebook reel URL to inspect.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Excel file to update.")
    parser.add_argument("--update-xlsx", action="store_true", help="Batch update reel_title/upload_date columns.")
    parser.add_argument("--limit", type=int, default=None, help="Maximum number of rows to fetch.")
    parser.add_argument("--start-row", type=int, default=2, help="Excel row number to start from.")
    parser.add_argument("--force", action="store_true", help="Overwrite existing reel_title/upload_date values.")
    parser.add_argument("--headful", action="store_true", help="Show the browser instead of running headless.")
    parser.add_argument("--save-every", type=int, default=10, help="Save after this many changed rows.")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()

    if args.update_xlsx:
        update_excel_reel_info(
            xlsx_path=args.xlsx,
            limit=args.limit,
            start_row=args.start_row,
            force=args.force,
            headless=not args.headful,
            save_every=args.save_every,
        )
        sys.exit(0)

    if not args.url:
        print("Usage: python fb_reel_title_browser.py <facebook_reel_url>")
        print("   or: python fb_reel_title_browser.py --update-xlsx --limit 20")
        sys.exit(1)

    try:
        title, date = get_reel_info(args.url, headless=not args.headful)
        print(f"Title: {title or 'Not found'}")
        print(f"Date: {date or 'Not found'}")
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)
