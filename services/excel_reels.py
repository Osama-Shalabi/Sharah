from __future__ import annotations

import hashlib
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


@dataclass(frozen=True)
class ExcelReelRow:
    reel_url: str
    thumbnail_url: str
    reel_title: str
    upload_date: str


def _norm(s: Any) -> str:
    return str(s or "").strip()


def _stable_id_from_url(url: str) -> str:
    return hashlib.sha1(url.encode("utf-8", errors="ignore")).hexdigest()


def _read_xlsx_rows(path: Path) -> List[ExcelReelRow]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return []

    cols = [_norm(c).lower() for c in header]
    try:
        url_idx = cols.index("reel_url")
    except ValueError:
        url_idx = None
    try:
        thumb_idx = cols.index("thumbnail_url")
    except ValueError:
        thumb_idx = None
    try:
        title_idx = cols.index("reel_title")
    except ValueError:
        title_idx = None
    try:
        date_idx = cols.index("upload_date")
    except ValueError:
        date_idx = None

    if url_idx is None:
        return []

    out: List[ExcelReelRow] = []
    for row in it:
        if not row:
            continue
        reel_url = _norm(row[url_idx] if url_idx < len(row) else "")
        if not reel_url:
            continue
        thumbnail_url = _norm(row[thumb_idx] if thumb_idx is not None and thumb_idx < len(row) else "")
        reel_title = _norm(row[title_idx] if title_idx is not None and title_idx < len(row) else "")
        upload_date = _norm(row[date_idx] if date_idx is not None and date_idx < len(row) else "")
        out.append(
            ExcelReelRow(
                reel_url=reel_url,
                thumbnail_url=thumbnail_url,
                reel_title=reel_title,
                upload_date=upload_date,
            )
        )
    return out


_CACHE: Dict[str, Tuple[float, List[ExcelReelRow]]] = {}


def list_reels_from_excel(
    *,
    xlsx_path: Path,
    limit: Optional[int] = 100,
    offset: int = 0,
) -> List[Dict[str, Any]]:
    p = Path(xlsx_path)
    if not p.exists() or not p.is_file():
        return []

    offset = max(0, int(offset or 0))
    limit_n = None if limit is None else max(1, int(limit))

    mtime = os.path.getmtime(p)
    key = str(p.resolve())
    cached = _CACHE.get(key)
    if not cached or cached[0] != mtime:
        _CACHE[key] = (mtime, _read_xlsx_rows(p))

    rows = _CACHE[key][1]
    sliced = rows[offset:] if limit_n is None else rows[offset : offset + limit_n]

    return [
        {
            "id": _stable_id_from_url(r.reel_url),
            "reel_url": r.reel_url,
            "thumbnail_url": r.thumbnail_url,
            "reel_title": r.reel_title,
            "upload_date": r.upload_date,
        }
        for r in sliced
    ]
